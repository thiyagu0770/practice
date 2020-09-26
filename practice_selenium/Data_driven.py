from  practice_selenium.excel_model import read
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By

# driver = webdriver.Chrome(r"C:\Users\ELCOT\PycharmProjects\Thiyagu\drivers/chromedriver.exe")
# driver.maximize_window()
# excel_location = r"D:\Selenium\Details.xlsx"
# x = openpyxl.load_workbook(excel_location)
# excel_location = r"D:\Selenium\Details1.xlsx"
# a = openpyxl.load_workbook(excel_location)
# b = a.create_sheet("data",1)
# c = b.cell(1, 1)
# c.value = "java"
# a.save(excel_location)











# excel_location = r"D:\Selenium\Company.xlsx"
# book = openpyxl.load_workbook(excel_location)
# x = book.active
# y = x.cell(1, 1)
# y.value = "abu"
# book.save(excel_location)
# print("Ok")











# driver = webdriver.Chrome(executable_path=r"C:\Users\ELCOT\PycharmProjects\Thiyagu\drivers\chromedriver.exe")
# driver.maximize_window()
# driver.get("http://adactinhotelapp.com/")
# excel_location = openpyxl.load_workbook(r"D:\Selenium\adactin.xlsx")
# adactin = excel_location.active
# rows = adactin.max_row
# column = adactin.max_column
# for r in range(2, rows+1):
#     for c in range(1, column+1):
#         user_name1 = driver.find_element(By.ID, "username")
#         user_name1.send_keys(adactin.cell(r, c).value)
#         break
#     for c in range(2, column+1):
#         pass_word1 = driver.find_element(By.ID, "password")
#         pass_word1.send_keys(adactin.cell(r, c).value)
#         break
#     login1 = driver.find_element(By.ID, "login").click()
#     time.sleep(3)
#     location1 = driver.find_element(By.ID,"location")
#     s_location1 = Select(location1).select_by_index(2)
#     driver.find_element(By.ID,"Submit").submit()
#     time.sleep(3)
#     btn1 = driver.find_element(By.ID,"radiobutton_1").click()
#     driver.find_element(By.ID,"continue").click()
#     for c in range(3,column+1):
#         First_name = driver.find_element(By.ID, "first_name")
#         First_name.send_keys(adactin.cell(r, c).value)
#         break
#     for c in range(4, column+1):
#         Last_name = driver.find_element(By.ID, "last_name")
#         Last_name.send_keys(adactin.cell(r, c).value)
#         break
#     for c in range(5, column+1):
#         address = driver.find_element(By.ID, "address")
#         address.send_keys(adactin.cell(r, c).value)
#         break
#     for c in range(6,column+1):
#         ccn = driver.find_element(By.ID, "cc_num")
#         ccn.send_keys(adactin.cell(r, c).value)
#         break
#     cc_type1 = driver.find_element(By.ID, "cc_type")
#     s_cc_type1 = Select(cc_type1).select_by_index(3)
#     month1 = driver.find_element(By.ID,"cc_exp_month")
#     s_month1 = Select(month1).select_by_index(4)
#     year1 = driver.find_element(By.ID,"cc_exp_year")
#     s_year1 = Select(year1).select_by_index(5)
#     for c in range(7, column+1):
#         cvv1 = driver.find_element(By.ID, "cc_cvv")
#         cvv1.send_keys(adactin.cell(r, c).value)
#         break
#     book_now1 = driver.find_element(By.ID, "book_now").click()
#     time.sleep(5)
#     order_no1 = driver.find_element(By.ID,"order_no")
#     time.sleep(10)
#     print(order_no1.get_attribute("value"))
#     print()
#     driver.find_element(By.ID,"logout").click()
#     time.sleep(5)
#     driver.find_element(By.XPATH,"//a[text()='Click here to login again']").click()






# user_name1 = driver.find_element(By.ID,"username")
# user_name1.send_keys(adactin.cell(3, 1).value)
# pass_word1 = driver.find_element(By.ID,"password")
# pass_word1.send_keys(adactin.cell(3, 2).value)
# login1 = driver.find_element(By.ID,"login").click()
# time.sleep(3)
# location1 = driver.find_element(By.ID,"location")
# s_location1 = Select(location1).select_by_index(2)
# driver.find_element(By.ID,"Submit").submit()
# time.sleep(3)
# btn1 = driver.find_element(By.ID,"radiobutton_1").click()
# driver.find_element(By.ID,"continue").click()
# First_name1 = driver.find_element(By.ID,"first_name")
# First_name1.send_keys(adactin.cell(3, 3).value)
# Last_name1 = driver.find_element(By.ID,"last_name")
# Last_name1.send_keys(adactin.cell(3, 4).value)
# address1 = driver.find_element(By.ID,"address")
# address1.send_keys(adactin.cell(3, 5).value)
# ccn1 = driver.find_element(By.ID,"cc_num")
# ccn1.send_keys(adactin.cell(3, 6).value)
# cc_type1 = driver.find_element(By.ID,"cc_type")
# s_cc_type1 = Select(cc_type1).select_by_index(3)
# month1 = driver.find_element(By.ID,"cc_exp_month")
# s_month1 = Select(month1).select_by_index(4)
# year1 = driver.find_element(By.ID,"cc_exp_year")
# s_year1 = Select(year1).select_by_index(5)
# cvv1 = driver.find_element(By.ID,"cc_cvv")
# cvv1.send_keys(adactin.cell(3, 7).value)
# book_now1 = driver.find_element(By.ID,"book_now").click()
# time.sleep(5)
# order_no1 = driver.find_element(By.ID,"order_no")
# time.sleep(10)
# print(order_no1.get_attribute("value"))
# print()
# driver.find_element(By.ID,"logout").click()
# time.sleep(5)
# driver.find_element(By.XPATH,"//a[text()='Click here to login again']").click()
#
# user_name2 = driver.find_element(By.ID,"username")
# user_name2.send_keys(adactin.cell(4, 1).value)
# pass_word2 = driver.find_element(By.ID,"password")
# pass_word2.send_keys(adactin.cell(4, 2).value)
# login2 = driver.find_element(By.ID,"login").click()
# time.sleep(3)
# location2 = driver.find_element(By.ID,"location")
# s_location2 = Select(location2).select_by_index(4)
# driver.find_element(By.ID,"Submit").submit()
# time.sleep(3)
# btn2 = driver.find_element(By.ID,"radiobutton_1").click()
# driver.find_element(By.ID,"continue").click()
# First_name2 = driver.find_element(By.ID,"first_name")
# First_name2.send_keys(adactin.cell(4, 3).value)
# Last_name2 = driver.find_element(By.ID,"last_name")
# Last_name2.send_keys(adactin.cell(4, 4).value)
# address2 = driver.find_element(By.ID,"address")
# address2.send_keys(adactin.cell(4, 5).value)
# ccn2 = driver.find_element(By.ID,"cc_num")
# ccn2.send_keys(adactin.cell(4, 6).value)
# cc_type2 = driver.find_element(By.ID,"cc_type")
# s_cc_type2 = Select(cc_type2).select_by_index(3)
# month2 = driver.find_element(By.ID,"cc_exp_month")
# s_month2 = Select(month2).select_by_index(3)
# year2 = driver.find_element(By.ID,"cc_exp_year")
# s_year2 = Select(year2).select_by_index(7)
# cvv2 = driver.find_element(By.ID,"cc_cvv")
# cvv2.send_keys(adactin.cell(4, 7).value)
# book_now2 = driver.find_element(By.ID,"book_now").click()
# time.sleep(5)
# order_no2 = driver.find_element(By.ID,"order_no")
# time.sleep(10)
# print(order_no2.get_attribute("value"))
# print()
# driver.find_element(By.ID,"logout").click()
# time.sleep(5)
# driver.find_element(By.XPATH,"//a[text()='Click here to login again']").click()
#
# user_name3 = driver.find_element(By.ID,"username")
# user_name3.send_keys(adactin.cell(5, 1).value)
# pass_word3 = driver.find_element(By.ID,"password")
# pass_word3.send_keys(adactin.cell(5, 2).value)
# login3 = driver.find_element(By.ID,"login").click()
# time.sleep(3)
# location3 = driver.find_element(By.ID,"location")
# s_location3 = Select(location3).select_by_index(4)
# driver.find_element(By.ID,"Submit").submit()
# time.sleep(3)
# btn3 = driver.find_element(By.ID,"radiobutton_1").click()
# driver.find_element(By.ID,"continue").click()
# First_name3 = driver.find_element(By.ID,"first_name")
# First_name3.send_keys(adactin.cell(5, 3).value)
# Last_name3 = driver.find_element(By.ID,"last_name")
# Last_name3.send_keys(adactin.cell(5, 4).value)
# address3 = driver.find_element(By.ID,"address")
# address3.send_keys(adactin.cell(5, 5).value)
# ccn3 = driver.find_element(By.ID,"cc_num")
# ccn3.send_keys(adactin.cell(5, 6).value)
# cc_type3 = driver.find_element(By.ID,"cc_type")
# s_cc_type3 = Select(cc_type3).select_by_index(3)
# month3 = driver.find_element(By.ID,"cc_exp_month")
# s_month3 = Select(month3).select_by_index(3)
# year3 = driver.find_element(By.ID,"cc_exp_year")
# s_year3 = Select(year3).select_by_index(7)
# cvv3 = driver.find_element(By.ID,"cc_cvv")
# cvv3.send_keys(adactin.cell(5, 7).value)
# book_now3 = driver.find_element(By.ID,"book_now").click()
# time.sleep(5)
# order_no3 = driver.find_element(By.ID,"order_no")
# time.sleep(10)
# print(order_no3.get_attribute("value"))
# print()
# driver.find_element(By.ID,"logout").click()
# time.sleep(5)
# driver.find_element(By.XPATH,"//a[text()='Click here to login again']").click()
#
# user_name4 = driver.find_element(By.ID,"username")
# user_name4.send_keys(adactin.cell(6, 1).value)
# pass_word4 = driver.find_element(By.ID,"password")
# pass_word4.send_keys(adactin.cell(6, 2).value)
# login4 = driver.find_element(By.ID,"login").click()
# time.sleep(3)
# location4 = driver.find_element(By.ID,"location")
# s_location4 = Select(location4).select_by_index(4)
# driver.find_element(By.ID,"Submit").submit()
# time.sleep(3)
# btn4 = driver.find_element(By.ID,"radiobutton_1").click()
# driver.find_element(By.ID,"continue").click()
# First_name4 = driver.find_element(By.ID,"first_name")
# First_name4.send_keys(adactin.cell(6, 3).value)
# Last_name4 = driver.find_element(By.ID,"last_name")
# Last_name4.send_keys(adactin.cell(6, 4).value)
# address4 = driver.find_element(By.ID,"address")
# address4.send_keys(adactin.cell(6, 5).value)
# ccn4 = driver.find_element(By.ID,"cc_num")
# ccn4.send_keys(adactin.cell(6, 6).value)
# cc_type4 = driver.find_element(By.ID,"cc_type")
# s_cc_type4 = Select(cc_type4).select_by_index(3)
# month4 = driver.find_element(By.ID,"cc_exp_month")
# s_month4 = Select(month4).select_by_index(3)
# year4 = driver.find_element(By.ID,"cc_exp_year")
# s_year4 = Select(year4).select_by_index(7)
# cvv4 = driver.find_element(By.ID,"cc_cvv")
# cvv4.send_keys(adactin.cell(6, 7).value)
# book_now4 = driver.find_element(By.ID,"book_now").click()
# time.sleep(5)
# order_no4 = driver.find_element(By.ID,"order_no")
# time.sleep(10)
# print(order_no4.get_attribute("value"))
# print()
# driver.find_element(By.ID,"logout").click()
# time.sleep(5)
# driver.find_element(By.XPATH,"//a[text()='Click here to login again']").click()

